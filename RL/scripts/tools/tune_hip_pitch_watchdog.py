"""Run hip-pitch step-response sims with a post-Done watchdog.

This helper is intentionally small and local to the tuning workflow.  It compares
IsaacLab sim output against real step-response data using the sign convention
that positive real angle equals negative sim angle.
"""

from __future__ import annotations

import argparse
import csv
import itertools
import math
import os
import queue
import shutil
import subprocess
import threading
import time
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(r"D:\Humanoid\Software\RL\Chaos_project\chaos01_train-main")
ROBO_ROOT = PROJECT_ROOT / "robolab" / "robolab-732b710f294926b8dd057057a6c3d86634c51e3f"
STEP_SCRIPT = ROBO_ROOT / "scripts" / "tools" / "step_response_sim.py"
REAL_XLSX = Path(r"D:\test_sim\excel\test_sim_0.xlsx")
REAL_CSV = Path(r"D:\test_sim\result\test_sim_pair_0_1_mean_by_target_timestep.csv")
PYTHON_EXE = Path(r"C:\Users\ngoc\AppData\Local\miniconda3\envs\env_isaaclab\python.exe")
OUT_ROOT = ROBO_ROOT / "step_response_data" / "hip_pitch_tune"


def real_targets() -> list[float]:
    if REAL_CSV.exists():
        with REAL_CSV.open(newline="") as f:
            return sorted({float(row["target_position"]) for row in csv.DictReader(f)})
    wb = openpyxl.load_workbook(REAL_XLSX, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col = {name: i for i, name in enumerate(header)}
    targets = sorted({float(row[col["target_position"]]) for row in ws.iter_rows(min_row=2, values_only=True)})
    return targets


def load_real(target: float) -> dict[float, tuple[float, float, float]]:
    if REAL_CSV.exists():
        rows: dict[float, tuple[float, float, float]] = {}
        with REAL_CSV.open(newline="") as f:
            for row in csv.DictReader(f):
                if float(row["target_position"]) != target:
                    continue
                t = float(row["time_ms"])
                real_angle = float(row["actual_angle_mean"])
                real_rpm = float(row["vel_mean"])
                actual_i = abs(float(row["actual_I_mean"]))
                real_torque = 0.3882 * actual_i + 0.2535
                rows[t] = (-real_angle, -real_rpm, real_torque)
        return rows

    wb = openpyxl.load_workbook(REAL_XLSX, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col = {name: i for i, name in enumerate(header)}
    rows: dict[float, tuple[float, float, float]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if float(row[col["target_position"]]) != target:
            continue
        t = float(row[col["time_ms"]])
        real_angle = float(row[col["actual_angle"]])
        real_rpm = float(row[col["actual_rpm"]])
        actual_i = abs(float(row[col["actual_I"]]))
        real_torque = 0.3882 * actual_i + 0.2535
        # User convention: positive real angle is negative sim angle.
        rows[t] = (-real_angle, -real_rpm, real_torque)
    return rows


def load_sim(out_dir: Path, sim_target: float) -> dict[float, tuple[float, float, float]]:
    stem = f"test_sim_{int(sim_target)}"
    candidates = [
        out_dir / f"{stem}_simulated.xlsx",
        out_dir / f"{stem}_nognd_simulated.xlsx",
        out_dir / f"{stem}_obs_simulated.xlsx",
        out_dir / f"{stem}_obs_nognd_simulated.xlsx",
    ]
    path = next((candidate for candidate in candidates if candidate.exists()), candidates[0])
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: dict[float, tuple[float, float, float]] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[1] is None:
            continue
        rows[float(row[1])] = (float(row[2]), float(row[3]), abs(float(row[4])))
    return rows


def score_one(real: dict[float, tuple[float, float, float]], sim: dict[float, tuple[float, float, float]]) -> dict[str, float]:
    common = sorted(set(real) & set(sim))
    if not common:
        raise RuntimeError("No common timestamps between real and sim data")

    def rmse(values):
        return math.sqrt(sum(v * v for v in values) / len(values))

    angle_errors = [sim[t][0] - real[t][0] for t in common]
    rpm_errors = [sim[t][1] - real[t][1] for t in common]
    torque_errors = [sim[t][2] - real[t][2] for t in common]
    return {
        "angle_rmse": rmse(angle_errors),
        "rpm_rmse": rmse(rpm_errors),
        "torque_rmse": rmse(torque_errors),
        "final_real": real[common[-1]][0],
        "final_sim": sim[common[-1]][0],
        "peak_rpm_real": max(abs(real[t][1]) for t in common),
        "peak_rpm_sim": max(abs(sim[t][1]) for t in common),
        "peak_torque_real": max(abs(real[t][2]) for t in common),
        "peak_torque_sim": max(abs(sim[t][2]) for t in common),
    }


def score_all(out_dir: Path, targets: list[float]) -> dict[str, float]:
    per_target = {}
    angle_sq = []
    rpm_sq = []
    torque_sq = []
    final_abs_errors = []
    peak_rpm_pairs = []
    peak_torque_pairs = []
    for target in targets:
        real = load_real(target)
        sim = load_sim(out_dir, -target)
        metrics = score_one(real, sim)
        per_target[target] = metrics
        common = sorted(set(real) & set(sim))
        angle_sq.extend((sim[t][0] - real[t][0]) ** 2 for t in common)
        rpm_sq.extend((sim[t][1] - real[t][1]) ** 2 for t in common)
        torque_sq.extend((sim[t][2] - real[t][2]) ** 2 for t in common)
        final_abs_errors.append(abs(metrics["final_sim"] - metrics["final_real"]))
        peak_rpm_pairs.append((metrics["peak_rpm_real"], metrics["peak_rpm_sim"]))
        peak_torque_pairs.append((metrics["peak_torque_real"], metrics["peak_torque_sim"]))

    def rmse(values):
        return math.sqrt(sum(values) / len(values))

    result = {
        "angle_rmse": rmse(angle_sq),
        "rpm_rmse": rmse(rpm_sq),
        "torque_rmse": rmse(torque_sq),
        "final_abs_mean": sum(final_abs_errors) / len(final_abs_errors),
        "peak_rpm_real_mean": sum(p[0] for p in peak_rpm_pairs) / len(peak_rpm_pairs),
        "peak_rpm_sim_mean": sum(p[1] for p in peak_rpm_pairs) / len(peak_rpm_pairs),
        "peak_torque_real_mean": sum(p[0] for p in peak_torque_pairs) / len(peak_torque_pairs),
        "peak_torque_sim_mean": sum(p[1] for p in peak_torque_pairs) / len(peak_torque_pairs),
    }
    torque_min_ratios = [
        sim_peak / max(1e-9, 1.25 * real_peak)
        for real_peak, sim_peak in peak_torque_pairs
    ]
    torque_deficits = [
        max(0.0, 1.25 * real_peak - sim_peak)
        for real_peak, sim_peak in peak_torque_pairs
    ]
    result["pos_vel_score"] = result["angle_rmse"] + 0.10 * result["rpm_rmse"] + 0.25 * result["final_abs_mean"]
    result["torque_gate_fail_count"] = sum(1 for deficit in torque_deficits if deficit > 1e-6)
    result["torque_gate_deficit_mean"] = sum(torque_deficits) / len(torque_deficits)
    result["torque_gate_min_ratio"] = min(torque_min_ratios)
    result["torque_stage_score"] = result["torque_rmse"] + result["torque_gate_deficit_mean"]
    result["weighted"] = result["pos_vel_score"]
    return result


def _reader_thread(stream, output_queue: queue.Queue[str]):
    for line in iter(stream.readline, ""):
        output_queue.put(line)
    stream.close()


def run_case(case_name: str, params: dict[str, float], targets: list[float], duration: float) -> dict[str, float]:
    sim_targets = [-target for target in targets]
    out_dir = OUT_ROOT / case_name
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    env = os.environ.copy()
    env.update(
        {
            "CHAOS_HIP_PITCH_STIFFNESS": str(params["kp"]),
            "CHAOS_HIP_PITCH_DAMPING": str(params["kd"]),
            "CHAOS_HIP_PITCH_ARMATURE": str(params["armature"]),
        }
    )
    for key in ("friction", "dynamic_friction", "viscous_friction"):
        if params.get(key) is not None:
            env[f"CHAOS_HIP_PITCH_{key.upper()}"] = str(params[key])

    cmd = [
        str(PYTHON_EXE),
        "-u",
        str(STEP_SCRIPT),
        "--headless",
        "--joint",
        "right_hip_pitch_joint",
        "--targets",
        *[f"{target:g}" for target in sim_targets],
        "--duration",
        str(duration),
        "--no-ground",
        "--out_dir",
        str(out_dir),
    ]
    print(f"\n[RUN] {case_name} {params}")
    proc = subprocess.Popen(
        cmd,
        cwd=PROJECT_ROOT,
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        bufsize=1,
    )

    done_at = None
    start = time.time()
    assert proc.stdout is not None
    output_queue: queue.Queue[str] = queue.Queue()
    reader = threading.Thread(target=_reader_thread, args=(proc.stdout, output_queue), daemon=True)
    reader.start()
    while True:
        try:
            line = output_queue.get(timeout=0.2)
            stripped = line.rstrip()
            is_sample = "|" in stripped and stripped[:8].strip().replace(".", "", 1).lstrip("-").isdigit()
            if (
                not is_sample
                or stripped.startswith("  Target")
                or stripped.startswith("===")
                or stripped.startswith("Done.")
                or "[WARN]" in stripped
                or "All joints" in stripped
                or "Joint      :" in stripped
                or "Limits" in stripped
            ):
                print(stripped)
            if "Done. Files:" in line and done_at is None:
                done_at = time.time()
        except queue.Empty:
            if proc.poll() is not None:
                break

        if done_at is not None and time.time() - done_at > 5.0 and proc.poll() is None:
            print("[WATCHDOG] Done printed, process still alive after 5s; killing.")
            proc.kill()
            break
        if time.time() - start > 900.0 and proc.poll() is None:
            proc.kill()
            raise TimeoutError(f"case {case_name} exceeded 900s")

    metrics = score_all(out_dir, targets)
    print(
        "[SCORE] pos_vel={pos_vel_score:.3f} angle_rmse={angle_rmse:.3f} rpm_rmse={rpm_rmse:.3f} "
        "final_abs_mean={final_abs_mean:.3f} torque_gate_fail={torque_gate_fail_count:.0f} "
        "torque_min_ratio={torque_gate_min_ratio:.2f} peak_rpm_mean(real/sim)={peak_rpm_real_mean:.1f}/{peak_rpm_sim_mean:.1f} "
        "peak_tau_mean(real/sim)={peak_torque_real_mean:.2f}/{peak_torque_sim_mean:.2f}".format(**metrics)
    )
    return metrics


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=float, default=None)
    parser.add_argument("--duration", type=float, default=3.0)
    parser.add_argument("--mode", choices=["baseline", "grid"], default="baseline")
    parser.add_argument("--angle-threshold", type=float, default=2.5)
    parser.add_argument("--rpm-threshold", type=float, default=2.5)
    args = parser.parse_args()

    targets = [args.target] if args.target is not None else real_targets()
    print("REAL_TARGETS", targets)

    if args.mode == "baseline":
        candidates = [
            {
                "kp": 40.0,
                "kd": 5.0,
                "armature": 0.10,
            }
        ]
    else:
        seeds = [
            (20.0, 2.0, 0.06),
            (30.0, 3.0, 0.06),
            (35.0, 3.5, 0.08),
            (40.0, 4.0, 0.08),
            (45.0, 4.5, 0.08),
            (50.0, 5.0, 0.08),
            (60.0, 6.0, 0.08),
            (90.0, 6.0, 0.08),
            (40.0, 5.0, 0.10),
            (50.0, 6.0, 0.10),
            (60.0, 7.0, 0.12),
        ]
        candidates = [
            {"kp": kp, "kd": kd, "armature": armature}
            for kp, kd, armature in seeds
        ]

    def selection_key(metrics: dict[str, float]) -> tuple[float, ...]:
        pos_vel_pass = metrics["angle_rmse"] <= args.angle_threshold and metrics["rpm_rmse"] <= args.rpm_threshold
        if pos_vel_pass:
            return (
                0.0,
                metrics["torque_gate_fail_count"],
                metrics["torque_gate_deficit_mean"],
                metrics["torque_stage_score"],
                metrics["pos_vel_score"],
            )
        return (
            1.0,
            metrics["pos_vel_score"],
            metrics["angle_rmse"],
            metrics["rpm_rmse"],
        )

    best = None
    for i, params in enumerate(candidates):
        metrics = run_case(f"case_{i:03d}", params, targets, args.duration)
        if best is None or selection_key(metrics) < selection_key(best[0]):
            best = (metrics, params)
            stage = "torque" if selection_key(metrics)[0] == 0.0 else "pos_vel"
            print(f"[BEST] case_{i:03d} {params} stage={stage} pos_vel={metrics['pos_vel_score']:.3f}")

    if best is not None:
        metrics, params = best
        print("\nBEST_PARAMS", params)
        print("BEST_METRICS", metrics)


if __name__ == "__main__":
    main()
